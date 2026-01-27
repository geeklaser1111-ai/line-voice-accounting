"""匯出 API"""
import io
import csv
from datetime import datetime
from fastapi import APIRouter, Request
from fastapi.responses import StreamingResponse
from typing import Optional
from database import get_all_transactions_for_export
from routers.auth import get_user_id_from_request

router = APIRouter(prefix="/api/export", tags=["匯出"])


@router.get("/csv")
async def export_csv(
    request: Request,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """匯出 CSV 檔案"""
    user_id = get_user_id_from_request(request)

    transactions = get_all_transactions_for_export(
        user_id=user_id,
        start_date=start_date,
        end_date=end_date
    )

    # 建立 CSV
    output = io.StringIO()
    writer = csv.writer(output)

    # 寫入標頭
    writer.writerow(["日期", "類型", "分類", "金額", "描述"])

    # 寫入資料
    for t in transactions:
        type_text = "收入" if t["type"] == "income" else "支出"
        writer.writerow([
            t["created_at"],
            type_text,
            t["category"],
            t["amount"],
            t["description"] or ""
        ])

    output.seek(0)

    # 產生檔名
    filename = f"accounting_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/excel")
async def export_excel(
    request: Request,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """匯出 Excel 檔案"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    user_id = get_user_id_from_request(request)

    transactions = get_all_transactions_for_export(
        user_id=user_id,
        start_date=start_date,
        end_date=end_date
    )

    # 建立工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "記帳明細"

    # 樣式設定
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # 寫入標頭
    headers = ["日期", "類型", "分類", "金額", "描述"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 寫入資料
    income_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    expense_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    for row, t in enumerate(transactions, 2):
        type_text = "收入" if t["type"] == "income" else "支出"
        fill = income_fill if t["type"] == "income" else expense_fill

        data = [
            t["created_at"],
            type_text,
            t["category"],
            t["amount"],
            t["description"] or ""
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.fill = fill
            if col == 4:  # 金額欄位靠右對齊
                cell.alignment = Alignment(horizontal="right")

    # 調整欄寬
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 30

    # 儲存到記憶體
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 產生檔名
    filename = f"accounting_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
